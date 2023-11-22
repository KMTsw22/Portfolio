import time
from openpyxl import Workbook
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium import webdriver

from openpyxl.utils import get_column_letter

class Market:
    def __init__(self, URL, title, start, end):
        try:
            self.error = False
            self.title = title
            self.start = start
            self.end = end
            self.DayList = []
            self.driver = webdriver.Chrome()

            self.driver.get(URL)

            self.table = self.driver.find_element(By.ID, 'footer')
            actions = ActionChains(self.driver)
            actions.move_to_element(self.table).perform()

            self.frame = self.driver.find_element(By.XPATH,"//iframe[@title='일별시세']")  # 프레임을 찾는 XPath 식
            self.driver.switch_to.frame(self.frame)
        except Exception as e:
            self.ReturnFalse()
            pass
    def FindStart(self):
        try:
            visited = True
            while True:
                time.sleep(0.1)
                self.driver.implicitly_wait(10)
                for i in range(1, 8):
                    day = self.driver.find_element(By.XPATH, f'/html/body/div/table/tbody/tr[{i}]/td[1]')
                    if day.text <= self.start:
                        visited = False
                        break
                if not visited:
                    return True
                time.sleep(0.1)
                self.driver.implicitly_wait(10)
                time.sleep(0.1)
                self.driver.find_element(By.CLASS_NAME, 'next').click()
                time.sleep(0.1)
        except Exception as e:
            return False
    def BeforEnd(self):
        if not self.FindStart():
            return False
        visited3 = True
        while True:
            time.sleep(0.1)
            for i in range(7, 0, -1):
                day = self.driver.find_element(By.XPATH, f'/html/body/div/table/tbody/tr[{i}]/td[1]')
                if day.text > self.end:
                    visited3 = False
                    break
                elif day.text < self.start:
                    pass
                else:
                    date = self.driver.find_element(By.XPATH, f'/html/body/div/table/tbody/tr[{i}]/td[1]').text
                    cost_finish = self.driver.find_element(By.XPATH, f'/html/body/div/table/tbody/tr[{i}]/td[2]').text
                    cost_ration = self.driver.find_element(By.XPATH, f'/html/body/div/table/tbody/tr[{i}]/td[3]')
                    cost_ration_final = ""
                    color = cost_ration.value_of_css_property('color')
                    if color == "rgba(0, 58, 206, 1)":
                        cost_ration_final = "-"+cost_ration.text
                    else:
                        cost_ration_final = cost_ration.text
                    cost_updown = self.driver.find_element(By.XPATH, f'/html/body/div/table/tbody/tr[{i}]/td[4]').text
                    # cost_start = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[4]').text
                    # cost_high = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[5]').text
                    # cost_low = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[6]').text
                    Dict_ = {"날짜": date, "종가": cost_finish, "전일 대비": cost_ration_final, "등락률": cost_updown,
                             }
                    self.DayList.append(Dict_)

                    paging_div = self.driver.find_element(By.CLASS_NAME, "Paging")
                    selected = paging_div.find_element(By.CSS_SELECTOR, "a.on")
                    now = int(selected.text)
                    if now == 1:
                        visited3 = False
            if not visited3:
                return True

            paging_div = self.driver.find_element(By.CLASS_NAME, "Paging")
            selected = paging_div.find_element(By.CSS_SELECTOR, "a.on")
            now = int(selected.text)
            if now > 3:
                self.driver.find_element(By.XPATH, '/ html / body / div / div / a[3]').click()
            elif now == 1:
                return False
            else:
                self.driver.find_element(By.XPATH,f'/html/body/div/div/a[{now-1}]').click()
    def SaveFile(self):
        if not self.BeforEnd():
            return False
        try:
            File = Workbook()
            sheet = File.active
            sheet.column_dimensions[get_column_letter(1)].width = 10
            sheet.column_dimensions[get_column_letter(3)].width = 10
            sheet['A1'] = '날짜'
            sheet['B1'] = '종가'
            sheet['C1'] = '전일 대비'
            sheet['D1'] = '등락률'

            for i in range(2, len(self.DayList) + 2):
                sheet.cell(i, 1, self.DayList[i - 2]['날짜'])
                sheet.cell(i, 2, self.DayList[i - 2]['종가'])
                sheet.cell(i, 3, self.DayList[i - 2]['전일 대비'])
                sheet.cell(i, 4, self.DayList[i - 2]['등락률'])
            File.save(f'{self.title}_{self.start[:4]}{self.start[5:7]}{self.start[8:]}_{self.end[:4]}{self.end[5:7]}{self.end[8:]}.xlsx')

            return True
        except Exception as e:
            return False
    def ReturnFalse(self):
        return False

