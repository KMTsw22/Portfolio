import time
from openpyxl import Workbook
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium import webdriver

from openpyxl.utils import get_column_letter

class Overseas:
    def __init__(self, URL, title, start, end):
        try:
            self.error = True
            self.title = title
            self.start = start
            self.end = end
            self.DayList = []
            self.driver = webdriver.Chrome()
            self.driver.get(URL)
            self.table = self.driver.find_element(By.ID, 'footer')
            actions = ActionChains(self.driver)
            actions.move_to_element(self.table).perform()
        except Exception as e:
            self.error = False
            return
    def FindStart(self):
        try:
            visited = True
            while True:
                time.sleep(0.1)
                self.driver.implicitly_wait(10)
                for i in range(1, 11):
                    day = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[1]')
                    if day.text <= self.start:
                        visited = False
                        break
                if not visited:
                    return True
                time.sleep(0.1)
                self.driver.implicitly_wait(10)
                actions = ActionChains(self.driver)
                actions.move_to_element(self.table).perform()
                time.sleep(0.1)
                self.driver.find_element(By.CLASS_NAME, 'next').click()
                time.sleep(0.1)
        except Exception as e:
            return False

    def BeforeEnd(self):
        try:
            self.FindStart()
            visited3 = True
            while True:
                time.sleep(0.1)
                for i in range(10, 0, -1):
                    day = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[1]')
                    if day.text > self.end:
                        visited3 = False
                        break
                    elif day.text < self.start:
                        pass
                    else:
                        date = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[1]').text
                        cost_finish = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[2]').text
                        cost_ration = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[3]/span')
                        cost_ration_final = ""
                        color = cost_ration.value_of_css_property('color')
                        if color == "rgba(0, 58, 206, 1)":
                            cost_ration_final = "-" + cost_ration.text
                        else:
                            cost_ration_final = cost_ration.text
                        cost_start = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[4]').text
                        cost_high = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[5]').text
                        cost_low = self.driver.find_element(By.XPATH, f'//*[@id="dayTable"]/tbody/tr[{i}]/td[6]').text
                        Dict_ = {"날짜": date, "종가": cost_finish, "전일 대비": cost_ration_final, "시가": cost_start,
                                 "고가": cost_high, "저가": cost_low}
                        self.DayList.append(Dict_)
                        paging_div = self.driver.find_element(By.ID, "dayPaging")
                        selected = paging_div.find_element(By.CSS_SELECTOR, "a.on")
                        if int(selected.text) == 1:
                            visited3 = False
                if not visited3:
                    return True
                paging_div = self.driver.find_element(By.ID, "dayPaging")
                selected = paging_div.find_element(By.CSS_SELECTOR, "a.on")
                now_index = int(selected.text)
                if now_index == 1:
                    return False
                elif now_index % 10 == 1:
                    self.driver.find_element(By.CLASS_NAME, 'prev').click()
                else:
                    self.driver.find_element(By.XPATH, f'//*[@id="dayLink{now_index - 1}"]').click()

        except Exception as e:
            return False
    def SaveFile(self):
        if not self.BeforeEnd():
            return False
        try:
            File = Workbook()
            sheet = File.active
            sheet.column_dimensions[get_column_letter(1)].width = 10
            sheet.column_dimensions[get_column_letter(6)].width = 10
            sheet['A1'] = '날짜'
            sheet['B1'] = '시가'
            sheet['C1'] = '종가'
            sheet['D1'] = '고가'
            sheet['E1'] = '저가'
            sheet['F1'] = '전일 대비'
            for i in range(2, len(self.DayList) + 2):
                sheet.cell(i, 1, self.DayList[i - 2]['날짜'])
                sheet.cell(i, 2, self.DayList[i - 2]['시가'])
                sheet.cell(i, 3, self.DayList[i - 2]['종가'])
                sheet.cell(i, 4, self.DayList[i - 2]['고가'])
                sheet.cell(i, 5, self.DayList[i - 2]['저가'])
                sheet.cell(i, 6, self.DayList[i - 2]['전일 대비'])

            File.save(f'{self.title}_{self.start[:4]}{self.start[5:7]}{self.start[8:]}_{self.end[:4]}{self.end[5:7]}{self.end[8:]}.xlsx')
            return True
        except Exception as e:
            return False
    def ReturnFalse(self):
        return False
